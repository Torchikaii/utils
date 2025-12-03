"""
BOM (Bill of Materials) Excel to Excel Converter with Consolidation

This script converts Excel BOM files to consolidated Excel format.
Consolidation groups identical components by Name or Manufacturer Part Number
and combines their designators with quantity counting.

Output: BOM_consolidated.xlsx
"""

from openpyxl import load_workbook, Workbook
from collections import defaultdict
import os
import ast

# Load source Excel file
excel_file = "BOM.xlsx"
if not os.path.exists(excel_file):
    print(f"Error: {excel_file} not found in current directory")
    exit(1)

# Read Excel headers and data
wb = load_workbook(excel_file)
ws = wb.active
headers = [cell.value for cell in ws[1]]

# Interactive column selection
print("Which columns you want to keep:")
recommended = [2, 3, 8, 12]  # Common Altium columns: description, designator, part_number, price
for i, header in enumerate(headers):
    tag = " (recommended)" if i in recommended else ""
    print(f"{i}: {header}{tag}")

# Get user preferences
selected = ast.literal_eval(input("Enter column indices as list (e.g., [2,3,8,12]): "))
currency = input("Which currency format did you exported in altium?: ")
consolidate = input("Do you want to consolidate parts? (y/n): ").lower() == 'y'

# Extract selected columns from source data
data = []
for row in ws.iter_rows(values_only=True):
    if any(cell is not None for cell in row):
        data.append([row[i] if i < len(row) else None for i in selected])

selected_headers = [headers[i] for i in selected]

# Create new Excel workbook for output
new_wb = Workbook()
new_ws = new_wb.active

if consolidate:
    """
    EXCEL CONSOLIDATION LOGIC:
    
    Same consolidation approach as markdown version but outputs to Excel:
    1. Group components by 'Name' or 'Manufacturer Part Number 1' (exact match)
    2. Combine designators for identical components
    3. Add quantity column and calculate total cost using 'Supplier Unit Price 1'
    4. Output consolidated data to new Excel file
    """
    
    # Find key columns by exact name matching for reliability
    name_idx = next((i for i, h in enumerate(selected_headers) if str(h) == 'Name'), None)
    part_num_idx = next((i for i, h in enumerate(selected_headers) if str(h) == 'Manufacturer Part Number 1'), None)
    designator_idx = next((i for i, h in enumerate(selected_headers) if str(h) == 'Designator'), None)
    
    # Validate required columns for consolidation
    if name_idx is None and part_num_idx is None:
        print("Error: Cannot consolidate - 'Name' or 'Manufacturer Part Number 1' column not selected")
        print("Required columns: 'Name' (preferred) or 'Manufacturer Part Number 1' (exact name match)")
        exit(1)
    if designator_idx is None:
        print("Error: Cannot consolidate - 'Designator' column not selected")
        print("Required column: 'Designator' (exact name match)")
        exit(1)
    
    # Choose consolidation key
    key_idx = name_idx if name_idx is not None else part_num_idx
    
    # Group identical components
    parts = defaultdict(lambda: {'designators': [], 'data': None, 'quantity': 0})
    
    for row in data[1:]:  # Skip header row
        if any(cell for cell in row):
            part_key = str(row[key_idx]) if row[key_idx] else 'N/A'
            designator = str(row[designator_idx]) if row[designator_idx] else ''
            
            parts[part_key]['designators'].append(designator)
            parts[part_key]['data'] = row
            parts[part_key]['quantity'] += 1
    
    # Write consolidated data to Excel
    new_ws.append(selected_headers + ["Quantity"])  # Header row with quantity column
    
    # Calculate total cost if price column exists (exact name match)
    total_cost = 0
    price_idx = next((i for i, h in enumerate(selected_headers) if str(h) == 'Supplier Unit Price 1'), None)
    
    # Write consolidated component rows
    for part_key, info in sorted(parts.items()):
        row = info['data'][:]  # Copy original data
        # Combine all designators for this component
        row[designator_idx] = ', '.join(sorted(info['designators']))
        
        # Calculate cost contribution
        if price_idx is not None and isinstance(row[price_idx], (int, float)):
            total_cost += row[price_idx] * info['quantity']
        
        # Add row with quantity
        new_ws.append(row + [info['quantity']])
    
    # Add total cost summary if price data available
    if price_idx is not None:
        new_ws.append([])  # Empty row separator
        new_ws.append([f"Total BOM Cost: {currency}{total_cost:.2f}"])
else:
    """
    DIRECT EXCEL CONVERSION:
    Simple column extraction without consolidation.
    Copies selected columns directly to new Excel file.
    """
    
    # Write header row
    new_ws.append(selected_headers)
    
    # Copy all data rows with selected columns
    for row in data[1:]:
        if any(cell for cell in row):
            new_ws.append(row)

# Save consolidated Excel file
new_wb.save("BOM_consolidated.xlsx")
print("BOM saved to BOM_consolidated.xlsx")