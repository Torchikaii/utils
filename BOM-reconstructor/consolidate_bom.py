"""
BOM (Bill of Materials) Excel to Markdown Converter

This script converts Excel BOM files to Markdown format with optional consolidation.
Consolidation groups identical components (same Name or Part Number) and combines
their designators into a single row with quantity count.

Features:
- Interactive column selection
- Currency formatting
- Component consolidation by Name or Manufacturer Part Number
- Total cost calculation
- Flexible column arrangement support
"""

from openpyxl import load_workbook
from collections import defaultdict
import os
import ast

# Load Excel file - expects BOM.xlsx in current directory
excel_file = "BOM.xlsx"
if not os.path.exists(excel_file):
    print(f"Error: {excel_file} not found in current directory")
    exit(1)

# Read Excel headers from first row
wb = load_workbook(excel_file)
ws = wb.active
headers = [cell.value for cell in ws[1]]

# Display available columns with recommendations
print("Which columns you want to keep:")
recommended = [2, 3, 8, 12]  # Common Altium BOM columns: description, designator, part_number, price
for i, header in enumerate(headers):
    tag = " (recommended)" if i in recommended else ""
    print(f"{i}: {header}{tag}")

# Get user preferences
selected = ast.literal_eval(input("Enter column indices as list (e.g., [2,3,8,12]): "))
currency = input("Which currency format did you export in Altium?(e.g. $): ")
consolidate = input("Do you want to consolidate parts? (y/n): ").lower() == 'y'

# Extract only selected columns from Excel data
data = []
for row in ws.iter_rows(values_only=True):
    if any(cell is not None for cell in row):
        data.append([row[i] if i < len(row) else None for i in selected])

# Get headers for selected columns
selected_headers = [headers[i] for i in selected]

if consolidate:
    """
    CONSOLIDATION LOGIC:
    
    1. Find key columns by exact name matching for reliability:
       - 'Name': Component name (preferred grouping key)
       - 'Manufacturer Part Number 1': Alternative grouping key if Name not available
       - 'Designator': Component reference designators (R1, C2, etc.)
       - 'Supplier Unit Price 1': Unit price for cost calculation
    
    2. Group components by Name or Part Number (components with identical
       specifications but different designators)
    
    3. Combine designators from grouped components into comma-separated list
    
    4. Calculate total quantity and cost for each component group
    """
    
    # Search for required columns by exact name matching for reliability
    name_idx = next((i for i, h in enumerate(selected_headers) if str(h) == 'Name'), None)
    part_num_idx = next((i for i, h in enumerate(selected_headers) if str(h) == 'Manufacturer Part Number 1'), None)
    designator_idx = next((i for i, h in enumerate(selected_headers) if str(h) == 'Designator'), None)
    
    # Validate required columns are present
    if name_idx is None and part_num_idx is None:
        print("Error: Cannot consolidate - 'Name' or 'Manufacturer Part Number 1' column not selected")
        print("Required columns: 'Name' (preferred) or 'Manufacturer Part Number 1'")
        exit(1)
    if designator_idx is None:
        print("Error: Cannot consolidate - 'Designator' column not selected")
        print("Required column: 'Designator' (exact name match)")
        exit(1)
    
    # Choose grouping key: prefer Name over Part Number
    key_idx = name_idx if name_idx is not None else part_num_idx
    
    # Dictionary to store grouped components
    # Key: component name/part number, Value: {designators: [], data: row, quantity: count}
    parts = defaultdict(lambda: {'designators': [], 'data': None, 'quantity': 0})
    
    # Process each data row (skip header)
    for row in data[1:]:
        if any(cell for cell in row):
            # Extract grouping key (component identifier)
            part_key = str(row[key_idx]) if row[key_idx] else 'N/A'
            # Extract designator (component reference)
            designator = str(row[designator_idx]) if row[designator_idx] else ''
            
            # Group components by part_key, collect designators
            parts[part_key]['designators'].append(designator)
            parts[part_key]['data'] = row  # Store component data
            parts[part_key]['quantity'] += 1  # Count occurrences
    
    # Generate consolidated markdown table
    markdown = "# BOM (Bill of Materials) - Consolidated\n\n"
    markdown += "| " + " | ".join(selected_headers + ["Quantity"]) + " |\n"
    markdown += "|" + "------------|" * (len(selected_headers) + 1) + "\n"
    
    # Find price column for cost calculation using exact column name
    total_cost = 0
    price_idx = next((i for i, h in enumerate(selected_headers) if str(h) == 'Supplier Unit Price 1'), None)
    
    # Output consolidated rows
    for part_key, info in sorted(parts.items()):
        row = info['data'][:]  # Copy original row data
        # Replace single designator with comma-separated list of all designators
        row[designator_idx] = ', '.join(sorted(info['designators']))
        
        # Format cells and calculate costs
        row_str = []
        for i, cell in enumerate(row):
            if isinstance(cell, (int, float)) and cell != 0:
                row_str.append(f"{currency}{cell:.2f}")
                # Add to total cost if this is the price column
                if price_idx is not None and i == price_idx:
                    total_cost += cell * info['quantity']
            else:
                row_str.append(str(cell) if cell else 'N/A')
        markdown += "| " + " | ".join(row_str + [str(info['quantity'])]) + " |\n"
    
    # Add total cost if price column exists
    if price_idx is not None:
        markdown += f"\n**Total BOM Cost: {currency}{total_cost:.2f}**\n"
else:
    """
    DIRECT CONVERSION MODE:
    Simple Excel to Markdown conversion without consolidation.
    Each Excel row becomes one Markdown table row with selected columns.
    """
    
    # Generate basic markdown table
    markdown = "# BOM (Bill of Materials)\n\n"
    markdown += "| " + " | ".join(selected_headers) + " |\n"
    markdown += "|" + "------------|" * len(selected_headers) + "\n"
    
    # Convert each data row to markdown (skip header)
    for row in data[1:]:
        if any(cell for cell in row):
            row_str = []
            for cell in row:
                # Format numeric values with currency symbol
                if isinstance(cell, (int, float)) and cell != 0:
                    row_str.append(f"{currency}{cell:.2f}")
                else:
                    row_str.append(str(cell) if cell else 'N/A')
            markdown += "| " + " | ".join(row_str) + " |\n"

# Write output file
with open("BOM_consolidated.md", "w", encoding="utf-8") as f:
    f.write(markdown)

print("BOM saved to BOM_consolidated.md")